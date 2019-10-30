#read links(clones/is cloned by) from issues and write the results in excel

#log in using API
from jira import JIRA
options = {'server':'https....'}
jira = JIRA(options, basic_auth=(config.USER, config.PASS)
#read the ticket from Excel which contains KEY-WORD
import openpyxl

wb = openpyxl.load_workbook("twolabels.xlsx")

ws = wb[wb.sheetnames[0]]

# go through all rows in a specific colum

for cell in ws['C']:
    issue_title_excel = cell.value
    if "KEY-WORD" in issue_title_excel:
        issue_key_excel = ws['B'.format(cell.row)]
        if issue_key_excel != "key":
            print(issue_key_excel)
            issue = jira.issue(issue_key_excel, fields = "key, summary, issuelinks")
            #check all issuelinks check the KEY of the ticket in Jira and store the link(issue-key and summary) of "clones" in a cell of the same row
            for issue_link in issue.fields.issuelinks:
                if issue_link.type == "Cloner" and hasattr(issue_link, "outwardIssue"):
                    original_issue_key = issue_link.outwardIssue
                    original_issue_title = issue_link.outwardIssue.fields.summary
                    ws['D{}'.format(cell.row)] = original_issue_key
                    ws['E{}'.format(cell.row)] = original_issue_title


#end the session
wb.save("twolabels2.xlsx")
#search the issue-key from link in Excel and write the issue-key and summary which contains KEY-WORD in the same row
